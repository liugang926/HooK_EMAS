"""
Batch Operations Service - Asset Batch Import/Export/Operations

Handles batch operations for assets:
- Batch import from Excel/CSV files
- Export assets to Excel files
- Batch receive, return, and transfer operations

Following .cursorrules: All business logic must be encapsulated in services/ directory.
"""
from django.db import transaction
from django.utils import timezone
from django.http import HttpResponse
from typing import Dict, Any, List, Optional, Tuple
from decimal import Decimal, InvalidOperation
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import logging

from .base import BaseService

logger = logging.getLogger(__name__)


class BatchImportExportService(BaseService):
    """
    Batch Import/Export Service
    
    Handles Excel file processing for asset batch operations.
    """
    
    # Field mapping for import: Excel header -> model field
    IMPORT_FIELD_MAPPING = {
        '资产编号': 'asset_code',
        '资产名称': 'name',
        '资产分类': 'category_name',
        '品牌': 'brand',
        '型号': 'model',
        '序列号': 'serial_number',
        '计量单位': 'unit',
        '数量': 'quantity',
        '取得方式': 'acquisition_method',
        '取得日期': 'acquisition_date',
        '原值': 'original_value',
        '净值': 'current_value',
        '使用部门': 'using_department_name',
        '使用人': 'using_user_name',
        '存放位置': 'location_name',
        '管理部门': 'manage_department_name',
        '资产管理员': 'manager_name',
        '保修到期日': 'warranty_expiry',
        '备注': 'remark',
    }
    
    # Export columns configuration
    EXPORT_COLUMNS = [
        ('asset_code', '资产编号', 15),
        ('name', '资产名称', 20),
        ('category_name', '资产分类', 15),
        ('brand', '品牌', 12),
        ('model', '型号', 15),
        ('serial_number', '序列号', 18),
        ('unit', '计量单位', 10),
        ('quantity', '数量', 8),
        ('status_display', '状态', 10),
        ('acquisition_method_display', '取得方式', 12),
        ('acquisition_date', '取得日期', 12),
        ('original_value', '原值', 12),
        ('current_value', '净值', 12),
        ('using_department_name', '使用部门', 15),
        ('using_user_name', '使用人', 12),
        ('location_name', '存放位置', 15),
        ('manage_department_name', '管理部门', 15),
        ('manager_name', '资产管理员', 12),
        ('warranty_expiry', '保修到期日', 12),
        ('remark', '备注', 25),
        ('created_at', '创建时间', 18),
    ]
    
    # Acquisition method mapping
    ACQUISITION_METHOD_MAP = {
        '采购': 'purchase',
        '租赁': 'lease',
        '赠予': 'gift',
        '调入': 'transfer',
        '自建': 'self_build',
        '其他': 'other',
    }
    
    @classmethod
    def generate_import_template(cls) -> HttpResponse:
        """
        Generate Excel import template
        
        Returns:
            HttpResponse with Excel file
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '资产导入模板'
        
        # Header style
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Required fields (marked with *)
        required_fields = ['资产名称*', '原值*']
        optional_fields = [
            '资产编号', '资产分类', '品牌', '型号', '序列号', '计量单位', '数量',
            '取得方式', '取得日期', '净值', '使用部门', '使用人', '存放位置',
            '管理部门', '资产管理员', '保修到期日', '备注'
        ]
        headers = required_fields + optional_fields
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Add sample data row
        sample_data = [
            '笔记本电脑', '8888.00', '', '办公设备', 'Dell', 'Latitude 5520', 
            'SN001234', '台', '1', '采购', '2024-01-15', '8888.00',
            '研发部', '张三', '3楼A区', '行政部', '李四', '2027-01-15', '新购置'
        ]
        for col, value in enumerate(sample_data, 1):
            ws.cell(row=2, column=col, value=value)
        
        # Add instructions sheet
        ws2 = wb.create_sheet('填写说明')
        instructions = [
            ['字段说明：'],
            ['资产名称*', '必填，资产的名称'],
            ['原值*', '必填，资产的原始价值（数字）'],
            ['资产编号', '可选，不填则自动生成'],
            ['资产分类', '可选，填写分类名称'],
            ['取得方式', '可选值：采购、租赁、赠予、调入、自建、其他'],
            ['取得日期', '格式：YYYY-MM-DD'],
            ['保修到期日', '格式：YYYY-MM-DD'],
            ['数量', '默认为1'],
            [''],
            ['注意事项：'],
            ['1. 带*的字段为必填字段'],
            ['2. 日期格式请使用 YYYY-MM-DD'],
            ['3. 金额字段请填写数字，不要带货币符号'],
            ['4. 使用部门、使用人、存放位置等字段请填写系统中已存在的名称'],
        ]
        for row_idx, row_data in enumerate(instructions, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws2.cell(row=row_idx, column=col_idx, value=value)
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 50
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="asset_import_template.xlsx"'
        return response
    
    @classmethod
    @transaction.atomic
    def import_assets(cls, file_obj, user) -> Dict[str, Any]:
        """
        Import assets from Excel file
        
        Args:
            file_obj: Uploaded Excel file
            user: Current user
            
        Returns:
            Dict with import results including success count, error list, etc.
        """
        from apps.assets.models import Asset, AssetCategory, AssetOperation
        from apps.organizations.models import Department, Location
        from apps.accounts.models import User
        
        company = cls.get_user_company(user)
        if not company:
            return {'success': False, 'message': '无法找到公司信息', 'success_count': 0, 'errors': []}
        
        try:
            wb = openpyxl.load_workbook(file_obj)
            ws = wb.active
        except Exception as e:
            return {'success': False, 'message': f'无法读取Excel文件: {str(e)}', 'success_count': 0, 'errors': []}
        
        # Get headers from first row
        headers = [cell.value.replace('*', '') if cell.value else '' for cell in ws[1]]
        
        # Map headers to fields
        field_map = {}
        for col_idx, header in enumerate(headers):
            if header in cls.IMPORT_FIELD_MAPPING:
                field_map[col_idx] = cls.IMPORT_FIELD_MAPPING[header]
        
        # Cache for lookups to avoid N+1 queries
        category_cache = {c.name: c for c in AssetCategory.objects.filter(company=company)}
        department_cache = {d.name: d for d in Department.objects.filter(company=company)}
        location_cache = {loc.name: loc for loc in Location.objects.filter(company=company)}
        user_cache = {}
        for u in User.objects.filter(memberships__company=company):
            user_cache[u.display_name] = u
            user_cache[u.username] = u
        
        results = {
            'success': True,
            'success_count': 0,
            'error_count': 0,
            'errors': [],
            'created_assets': []
        }
        
        # Process each row (skip header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not any(row):
                continue
            
            row_data = {}
            for col_idx, value in enumerate(row):
                if col_idx in field_map:
                    row_data[field_map[col_idx]] = value
            
            # Validate required fields
            if not row_data.get('name'):
                results['errors'].append({'row': row_idx, 'error': '资产名称为必填项'})
                results['error_count'] += 1
                continue
            
            if not row_data.get('original_value'):
                results['errors'].append({'row': row_idx, 'error': '原值为必填项'})
                results['error_count'] += 1
                continue
            
            try:
                # Process asset data
                asset_data = cls._process_import_row(
                    row_data, company, category_cache, department_cache, 
                    location_cache, user_cache
                )
                
                # Generate asset code if not provided
                if not asset_data.get('asset_code'):
                    asset_data['asset_code'] = cls.generate_order_no('ZC')
                
                # Check for duplicate asset_code
                if Asset.objects.filter(company=company, asset_code=asset_data['asset_code']).exists():
                    results['errors'].append({
                        'row': row_idx, 
                        'error': f'资产编号 {asset_data["asset_code"]} 已存在'
                    })
                    results['error_count'] += 1
                    continue
                
                # Create asset
                asset_data['company'] = company
                asset_data['created_by'] = user
                
                asset = Asset.objects.create(**asset_data)
                
                # Record operation
                AssetOperation.objects.create(
                    asset=asset,
                    operation_type=AssetOperation.OperationType.CREATE,
                    description=f'批量导入创建资产: {asset.name}',
                    new_data={
                        'asset_code': asset.asset_code,
                        'name': asset.name,
                        'original_value': str(asset.original_value),
                    },
                    operator=user
                )
                
                results['success_count'] += 1
                results['created_assets'].append({
                    'id': asset.id,
                    'asset_code': asset.asset_code,
                    'name': asset.name
                })
                
            except Exception as e:
                logger.error(f'Import error at row {row_idx}: {str(e)}')
                results['errors'].append({'row': row_idx, 'error': str(e)})
                results['error_count'] += 1
        
        return results
    
    @classmethod
    def _process_import_row(
        cls, row_data: Dict, company, category_cache, department_cache, 
        location_cache, user_cache
    ) -> Dict:
        """Process a single import row and convert to asset data"""
        
        asset_data = {}
        
        # Basic fields
        for field in ['asset_code', 'name', 'brand', 'model', 'serial_number', 'unit', 'remark']:
            if row_data.get(field):
                asset_data[field] = str(row_data[field]).strip()
        
        # Quantity
        if row_data.get('quantity'):
            try:
                asset_data['quantity'] = int(row_data['quantity'])
            except (ValueError, TypeError):
                asset_data['quantity'] = 1
        
        # Numeric fields
        for field in ['original_value', 'current_value']:
            if row_data.get(field):
                try:
                    asset_data[field] = Decimal(str(row_data[field]).replace(',', ''))
                except (InvalidOperation, ValueError):
                    if field == 'original_value':
                        raise ValueError(f'原值格式不正确: {row_data[field]}')
        
        # Default current_value to original_value
        if 'original_value' in asset_data and 'current_value' not in asset_data:
            asset_data['current_value'] = asset_data['original_value']
        
        # Date fields
        for field in ['acquisition_date', 'warranty_expiry']:
            if row_data.get(field):
                try:
                    if hasattr(row_data[field], 'strftime'):
                        asset_data[field] = row_data[field]
                    else:
                        from datetime import datetime
                        date_str = str(row_data[field]).strip()
                        for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d']:
                            try:
                                asset_data[field] = datetime.strptime(date_str, fmt).date()
                                break
                            except ValueError:
                                continue
                except Exception:
                    pass  # Skip invalid dates
        
        # Acquisition method
        if row_data.get('acquisition_method'):
            method = row_data['acquisition_method']
            if method in cls.ACQUISITION_METHOD_MAP:
                asset_data['acquisition_method'] = cls.ACQUISITION_METHOD_MAP[method]
        
        # Category lookup
        if row_data.get('category_name'):
            category = category_cache.get(row_data['category_name'])
            if category:
                asset_data['category'] = category
        
        # Department lookup
        if row_data.get('using_department_name'):
            dept = department_cache.get(row_data['using_department_name'])
            if dept:
                asset_data['using_department'] = dept
        
        if row_data.get('manage_department_name'):
            dept = department_cache.get(row_data['manage_department_name'])
            if dept:
                asset_data['manage_department'] = dept
        
        # Location lookup
        if row_data.get('location_name'):
            loc = location_cache.get(row_data['location_name'])
            if loc:
                asset_data['location'] = loc
        
        # User lookup
        if row_data.get('using_user_name'):
            user = user_cache.get(row_data['using_user_name'])
            if user:
                asset_data['using_user'] = user
        
        if row_data.get('manager_name'):
            user = user_cache.get(row_data['manager_name'])
            if user:
                asset_data['manager'] = user
        
        return asset_data
    
    @classmethod
    def export_assets(cls, queryset, export_fields: List[str] = None) -> HttpResponse:
        """
        Export assets to Excel file
        
        Args:
            queryset: Asset queryset to export
            export_fields: Optional list of field names to export
            
        Returns:
            HttpResponse with Excel file
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '资产列表'
        
        # Determine columns to export
        columns = cls.EXPORT_COLUMNS
        if export_fields:
            columns = [(f, h, w) for f, h, w in cls.EXPORT_COLUMNS if f in export_fields]
        
        # Header style
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col_idx, (field, header, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Write data rows
        for row_idx, asset in enumerate(queryset, 2):
            for col_idx, (field, header, width) in enumerate(columns, 1):
                value = cls._get_export_value(asset, field)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="assets_export_{timestamp}.xlsx"'
        return response
    
    @classmethod
    def _get_export_value(cls, asset, field: str):
        """Get the value of a field for export"""
        
        # Handle display fields
        if field.endswith('_display'):
            method_name = f'get_{field[:-8]}_display'
            if hasattr(asset, method_name):
                return getattr(asset, method_name)()
            return ''
        
        # Handle related fields
        if field.endswith('_name'):
            related_field = field[:-5]
            related_obj = getattr(asset, related_field, None)
            if related_obj:
                if hasattr(related_obj, 'display_name'):
                    return related_obj.display_name
                elif hasattr(related_obj, 'name'):
                    return related_obj.name
                elif hasattr(related_obj, 'full_name'):
                    return related_obj.full_name
            return ''
        
        # Handle datetime fields
        if field in ['created_at', 'updated_at']:
            value = getattr(asset, field, None)
            if value:
                return value.strftime('%Y-%m-%d %H:%M:%S')
            return ''
        
        # Handle date fields
        if field in ['acquisition_date', 'warranty_expiry']:
            value = getattr(asset, field, None)
            if value:
                return value.strftime('%Y-%m-%d')
            return ''
        
        # Handle decimal fields
        if field in ['original_value', 'current_value', 'accumulated_depreciation']:
            value = getattr(asset, field, None)
            if value is not None:
                return float(value)
            return ''
        
        # Default field value
        return getattr(asset, field, '') or ''


class BatchOperationService(BaseService):
    """
    Batch Operation Service
    
    Handles batch receive, return, and transfer operations.
    """
    
    @classmethod
    @transaction.atomic
    def batch_receive(
        cls,
        asset_ids: List[int],
        receive_user_id: int,
        receive_department_id: int,
        receive_location_id: int,
        receive_date: str,
        reason: str,
        user,
        company_id: int = None
    ) -> Dict[str, Any]:
        """
        Batch receive assets
        
        Assigns multiple assets to a user/department/location.
        
        Args:
            asset_ids: List of asset IDs to receive
            receive_user_id: User ID receiving the assets
            receive_department_id: Department ID receiving the assets
            receive_location_id: Location ID for the assets (optional)
            receive_date: Date of receive
            reason: Reason for receiving
            user: Current user performing the operation
            company_id: Company ID from frontend (optional, uses user's company or first company if not provided)
            
        Returns:
            Dict with operation results
        """
        from apps.assets.models import Asset, AssetReceive, AssetReceiveItem, AssetOperation
        from apps.accounts.models import User as UserModel
        from apps.organizations.models import Department, Location, Company
        
        # Validate assets - first get by ID, then check company access
        assets = Asset.objects.filter(id__in=asset_ids, is_deleted=False)
        if not assets.exists():
            return {'success': False, 'message': '未找到有效的资产'}
        
        # Get company - prioritize frontend's selected company, then user's company, then first company
        company = None
        if company_id:
            try:
                company = Company.objects.get(id=company_id)
            except Company.DoesNotExist:
                pass
        if not company:
            company = cls.get_user_company(user)
        
        # Check if assets are available for receive
        invalid_assets = []
        for asset in assets:
            if asset.status not in [Asset.Status.IDLE]:
                invalid_assets.append(f'{asset.name}({asset.asset_code}): 状态为{asset.get_status_display()}')
        
        if invalid_assets:
            return {
                'success': False, 
                'message': '以下资产状态不允许领用',
                'invalid_assets': invalid_assets
            }
        
        # Get receive user, department, and location
        receive_user = UserModel.objects.get(id=receive_user_id)
        receive_department = Department.objects.get(id=receive_department_id) if receive_department_id else None
        receive_location = Location.objects.get(id=receive_location_id) if receive_location_id else None
        
        # Create receive record
        receive_no = cls.generate_order_no('LY')
        receive = AssetReceive.objects.create(
            receive_no=receive_no,
            company=company,
            status=AssetReceive.Status.COMPLETED,
            receive_user=receive_user,
            receive_department=receive_department,
            receive_date=receive_date,
            reason=reason,
            created_by=user
        )
        
        # Process each asset
        for asset in assets:
            # Create receive item
            AssetReceiveItem.objects.create(
                receive=receive,
                asset=asset,
                quantity=1
            )
            
            # Update asset status
            old_data = {
                'status': asset.get_status_display(),
                'using_user': asset.using_user.display_name if asset.using_user else None,
                'using_department': asset.using_department.name if asset.using_department else None,
                'location': asset.location.name if asset.location else None,
            }
            
            asset.status = Asset.Status.IN_USE
            asset.using_user = receive_user
            asset.using_department = receive_department
            if receive_location:
                asset.location = receive_location
            asset.save()
            
            # Record operation
            AssetOperation.objects.create(
                asset=asset,
                operation_type=AssetOperation.OperationType.RECEIVE,
                operation_no=receive_no,
                description=f'批量领用: {receive_user.display_name}',
                old_data=old_data,
                new_data={
                    'status': asset.get_status_display(),
                    'using_user': receive_user.display_name,
                    'using_department': receive_department.name if receive_department else None,
                    'location': receive_location.name if receive_location else asset.location.name if asset.location else None,
                },
                operator=user
            )
        
        return {
            'success': True,
            'receive_no': receive_no,
            'count': assets.count(),
            'message': f'成功领用 {assets.count()} 项资产'
        }
    
    @classmethod
    @transaction.atomic
    def batch_return(
        cls,
        asset_ids: List[int],
        return_date: str,
        reason: str,
        user
    ) -> Dict[str, Any]:
        """
        Batch return assets
        
        Returns multiple assets to idle status.
        
        Args:
            asset_ids: List of asset IDs to return
            return_date: Date of return
            reason: Reason for returning
            user: Current user performing the operation
            
        Returns:
            Dict with operation results
        """
        from apps.assets.models import Asset, AssetOperation
        
        # Validate assets - first get by ID, then check status
        assets = Asset.objects.filter(id__in=asset_ids, is_deleted=False)
        if not assets.exists():
            return {'success': False, 'message': '未找到有效的资产'}
        
        # Check if assets can be returned
        invalid_assets = []
        for asset in assets:
            if asset.status not in [Asset.Status.IN_USE, Asset.Status.BORROWED]:
                invalid_assets.append(f'{asset.name}({asset.asset_code}): 状态为{asset.get_status_display()}')
        
        if invalid_assets:
            return {
                'success': False, 
                'message': '以下资产状态不允许退还',
                'invalid_assets': invalid_assets
            }
        
        return_no = cls.generate_order_no('TH')
        
        # Process each asset
        for asset in assets:
            old_data = {
                'status': asset.get_status_display(),
                'using_user': asset.using_user.display_name if asset.using_user else None,
                'using_department': asset.using_department.name if asset.using_department else None,
            }
            
            # Update asset status
            asset.status = Asset.Status.IDLE
            asset.using_user = None
            asset.using_department = None
            asset.save()
            
            # Record operation
            AssetOperation.objects.create(
                asset=asset,
                operation_type=AssetOperation.OperationType.RETURN,
                operation_no=return_no,
                description=f'批量退还: {reason or "无说明"}',
                old_data=old_data,
                new_data={
                    'status': asset.get_status_display(),
                    'using_user': None,
                    'using_department': None,
                },
                operator=user
            )
        
        return {
            'success': True,
            'return_no': return_no,
            'count': assets.count(),
            'message': f'成功退还 {assets.count()} 项资产'
        }
    
    @classmethod
    @transaction.atomic
    def batch_transfer(
        cls,
        asset_ids: List[int],
        to_department_id: int,
        to_user_id: int,
        to_location_id: int,
        transfer_date: str,
        reason: str,
        user,
        company_id: int = None
    ) -> Dict[str, Any]:
        """
        Batch transfer assets
        
        Transfers multiple assets to a new department/user/location.
        
        Args:
            asset_ids: List of asset IDs to transfer
            to_department_id: Target department ID (optional)
            to_user_id: Target user ID (optional)
            to_location_id: Target location ID (optional)
            transfer_date: Date of transfer
            reason: Reason for transfer
            user: Current user performing the operation
            company_id: Company ID from frontend (optional)
            
        Returns:
            Dict with operation results
        """
        from apps.assets.models import Asset, AssetTransfer, AssetTransferItem, AssetOperation
        from apps.accounts.models import User as UserModel
        from apps.organizations.models import Department, Location, Company
        
        # At least one target must be specified
        if not any([to_department_id, to_user_id, to_location_id]):
            return {'success': False, 'message': '请至少指定一个调拨目标（部门、人员或位置）'}
        
        # Validate assets - first get by ID
        assets = Asset.objects.filter(id__in=asset_ids, is_deleted=False)
        if not assets.exists():
            return {'success': False, 'message': '未找到有效的资产'}
        
        # Get company - prioritize frontend's selected company, then user's company, then first company
        company = None
        if company_id:
            try:
                company = Company.objects.get(id=company_id)
            except Company.DoesNotExist:
                pass
        if not company:
            company = cls.get_user_company(user)
        
        # Get target objects
        to_department = Department.objects.get(id=to_department_id) if to_department_id else None
        to_user = UserModel.objects.get(id=to_user_id) if to_user_id else None
        to_location = Location.objects.get(id=to_location_id) if to_location_id else None
        
        # Get from department (from first asset)
        first_asset = assets.first()
        from_department = first_asset.using_department or first_asset.manage_department
        
        # Create transfer record
        transfer_no = cls.generate_order_no('DB')
        transfer = AssetTransfer.objects.create(
            transfer_no=transfer_no,
            company=company,
            status=AssetTransfer.Status.COMPLETED,
            from_department=from_department,
            to_department=to_department,
            to_user=to_user,
            to_location=to_location,
            transfer_date=transfer_date,
            reason=reason,
            created_by=user
        )
        
        # Process each asset
        for asset in assets:
            # Create transfer item
            AssetTransferItem.objects.create(
                transfer=transfer,
                asset=asset,
                quantity=1,
                from_user=asset.using_user,
                from_department=asset.using_department,
                from_location=asset.location
            )
            
            # Record old data
            old_data = {
                'using_user': asset.using_user.display_name if asset.using_user else None,
                'using_department': asset.using_department.name if asset.using_department else None,
                'location': asset.location.name if asset.location else None,
            }
            
            # Update asset
            if to_department:
                asset.using_department = to_department
            if to_user:
                asset.using_user = to_user
                asset.status = Asset.Status.IN_USE
            if to_location:
                asset.location = to_location
            asset.save()
            
            # Record operation
            changes = []
            if to_department:
                changes.append(f'部门变更为 {to_department.name}')
            if to_user:
                changes.append(f'使用人变更为 {to_user.display_name}')
            if to_location:
                changes.append(f'位置变更为 {to_location.name}')
            
            AssetOperation.objects.create(
                asset=asset,
                operation_type=AssetOperation.OperationType.TRANSFER,
                operation_no=transfer_no,
                description=f'批量调拨: {"; ".join(changes)}',
                old_data=old_data,
                new_data={
                    'using_user': asset.using_user.display_name if asset.using_user else None,
                    'using_department': asset.using_department.name if asset.using_department else None,
                    'location': asset.location.name if asset.location else None,
                },
                operator=user
            )
        
        return {
            'success': True,
            'transfer_no': transfer_no,
            'count': assets.count(),
            'message': f'成功调拨 {assets.count()} 项资产'
        }
    
    @classmethod
    @transaction.atomic
    def batch_delete(
        cls,
        asset_ids: List[int],
        user
    ) -> Dict[str, Any]:
        """
        Batch soft delete assets
        
        Args:
            asset_ids: List of asset IDs to delete
            user: Current user performing the operation
            
        Returns:
            Dict with operation results
        """
        from apps.assets.models import Asset
        
        # Validate assets - first get by ID
        assets = Asset.objects.filter(id__in=asset_ids, is_deleted=False)
        if not assets.exists():
            return {'success': False, 'message': '未找到有效的资产'}
        
        # Check if assets can be deleted
        invalid_assets = []
        for asset in assets:
            if asset.status in [Asset.Status.IN_USE, Asset.Status.BORROWED, Asset.Status.MAINTENANCE]:
                invalid_assets.append(f'{asset.name}({asset.asset_code}): 状态为{asset.get_status_display()}')
        
        if invalid_assets:
            return {
                'success': False, 
                'message': '以下资产状态不允许删除（请先退还或归还）',
                'invalid_assets': invalid_assets
            }
        
        # Soft delete assets
        count = assets.update(is_deleted=True, deleted_at=timezone.now())
        
        return {
            'success': True,
            'count': count,
            'message': f'成功删除 {count} 项资产'
        }
